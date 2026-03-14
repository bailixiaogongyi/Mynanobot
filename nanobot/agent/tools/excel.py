"""Excel operations tools using openpyxl."""

import json
from pathlib import Path
from typing import Any

from nanobot.agent.tools.base import Tool


def _resolve_path(path: str, workspace: Path | None = None, allowed_dir: Path | None = None) -> Path:
    """Resolve path against workspace and enforce directory restriction."""
    p = Path(path).expanduser()
    if not p.is_absolute() and workspace:
        p = workspace / p
    resolved = p.resolve()
    if allowed_dir:
        try:
            resolved.relative_to(allowed_dir.resolve())
        except ValueError:
            raise PermissionError(f"Path {path} is outside allowed directory {allowed_dir}")
    return resolved


class ExcelReadTool(Tool):
    """Tool to read data from Excel file."""

    def __init__(self, workspace: Path | None = None, allowed_dir: Path | None = None):
        self._workspace = workspace
        self._allowed_dir = allowed_dir

    @property
    def name(self) -> str:
        return "excel_read"

    @property
    def description(self) -> str:
        return "Read data from Excel file. Returns data in JSON format with sheet names and row data."

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "The absolute path to the Excel file (.xlsx)"
                },
                "sheet_name": {
                    "type": "string",
                    "description": "Sheet name to read. Default: first sheet"
                },
                "header_row": {
                    "type": "boolean",
                    "description": "Whether first row is header. Default: true"
                },
                "max_rows": {
                    "type": "integer",
                    "description": "Maximum number of rows to read. Default: all"
                }
            },
            "required": ["path"]
        }

    async def execute(self, path: str, sheet_name: str | None = None,
                     header_row: bool = True, max_rows: int | None = None, **kwargs: Any) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return "Error: openpyxl is required. Install with: pip install openpyxl"

        try:
            file_path = _resolve_path(path, self._workspace, self._allowed_dir)
            if not file_path.exists():
                return f"Error: File not found: {path}"
            if file_path.suffix.lower() not in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
                return f"Error: Not an Excel file: {path}"

            wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)

            if sheet_name:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    return f"Error: Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}"
            else:
                ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            if max_rows:
                rows = rows[:max_rows]

            headers = []
            data = []

            if header_row and rows:
                headers = [str(cell) if cell is not None else "" for cell in rows[0]]
                for row in rows[1:]:
                    row_data = {headers[i]: str(cell) if cell is not None else "" for i, cell in enumerate(row)}
                    data.append(row_data)
            else:
                for row in rows:
                    data.append([str(cell) if cell is not None else "" for cell in row])

            wb.close()

            result = {
                "sheet_name": ws.title,
                "total_rows": len(data),
                "headers": headers if header_row else None,
                "data": data
            }

            return json.dumps(result, ensure_ascii=False, indent=2)

        except PermissionError as e:
            return f"Error: {e}"
        except Exception as e:
            return f"Error reading Excel: {str(e)}"


class ExcelWriteTool(Tool):
    """Tool to write data to Excel file."""

    def __init__(self, workspace: Path | None = None, allowed_dir: Path | None = None):
        self._workspace = workspace
        self._allowed_dir = allowed_dir

    @property
    def name(self) -> str:
        return "excel_write"

    @property
    def description(self) -> str:
        return "Write data to Excel file. Supports creating new file or appending to existing."

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "The absolute path for the output Excel file"
                },
                "data": {
                    "type": "object",
                    "description": "Data to write as JSON object with 'headers' and 'rows' keys"
                },
                "sheet_name": {
                    "type": "string",
                    "description": "Sheet name. Default: 'Sheet1'"
                },
                "append": {
                    "type": "boolean",
                    "description": "Append to existing file. Default: false"
                }
            },
            "required": ["path", "data"]
        }

    async def execute(self, path: str, data: dict[str, Any], sheet_name: str = "Sheet1",
                     append: bool = False, **kwargs: Any) -> str:
        try:
            from openpyxl import load_workbook, Workbook
        except ImportError:
            return "Error: openpyxl is required. Install with: pip install openpyxl"

        try:
            file_path = _resolve_path(path, self._workspace, self._allowed_dir)

            if append and file_path.exists():
                wb = load_workbook(filename=str(file_path))
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.create_sheet(sheet_name)
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = sheet_name

            headers = data.get("headers", [])
            rows = data.get("rows", [])

            if headers:
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

            start_row = len(headers) + 1 if headers else 1
            for row_idx, row_data in enumerate(rows, start_row):
                if isinstance(row_data, dict):
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col, value=row_data.get(header, ""))
                else:
                    for col, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col, value=value)

            file_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(file_path))
            wb.close()

            return f"Successfully wrote {len(rows)} rows to {file_path}"

        except PermissionError as e:
            return f"Error: {e}"
        except Exception as e:
            return f"Error writing Excel: {str(e)}"


class ExcelListSheetsTool(Tool):
    """Tool to list all sheets in Excel file."""

    def __init__(self, workspace: Path | None = None, allowed_dir: Path | None = None):
        self._workspace = workspace
        self._allowed_dir = allowed_dir

    @property
    def name(self) -> str:
        return "excel_list_sheets"

    @property
    def description(self) -> str:
        return "List all sheet names in an Excel file."

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "The absolute path to the Excel file"
                }
            },
            "required": ["path"]
        }

    async def execute(self, path: str, **kwargs: Any) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return "Error: openpyxl is required. Install with: pip install openpyxl"

        try:
            file_path = _resolve_path(path, self._workspace, self._allowed_dir)
            if not file_path.exists():
                return f"Error: File not found: {path}"

            wb = load_workbook(filename=str(file_path), read_only=True)
            sheets = wb.sheetnames
            wb.close()

            result = {
                "file": str(file_path),
                "sheet_count": len(sheets),
                "sheets": sheets
            }

            return json.dumps(result, ensure_ascii=False, indent=2)

        except PermissionError as e:
            return f"Error: {e}"
        except Exception as e:
            return f"Error listing sheets: {str(e)}"


class ExcelCreateSheetTool(Tool):
    """Tool to create a new sheet in Excel file."""

    def __init__(self, workspace: Path | None = None, allowed_dir: Path | None = None):
        self._workspace = workspace
        self._allowed_dir = allowed_dir

    @property
    def name(self) -> str:
        return "excel_create_sheet"

    @property
    def description(self) -> str:
        return "Create a new sheet in an existing Excel file."

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "The absolute path to the Excel file"
                },
                "sheet_name": {
                    "type": "string",
                    "description": "Name for the new sheet"
                }
            },
            "required": ["path", "sheet_name"]
        }

    async def execute(self, path: str, sheet_name: str, **kwargs: Any) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return "Error: openpyxl is required. Install with: pip install openpyxl"

        try:
            file_path = _resolve_path(path, self._workspace, self._allowed_dir)

            if file_path.exists():
                wb = load_workbook(filename=str(file_path))
            else:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Sheet1"

            if sheet_name in wb.sheetnames:
                wb.close()
                return f"Error: Sheet '{sheet_name}' already exists"

            wb.create_sheet(sheet_name)
            file_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(file_path))
            wb.close()

            return f"Successfully created sheet '{sheet_name}' in {file_path}"

        except PermissionError as e:
            return f"Error: {e}"
        except Exception as e:
            return f"Error creating sheet: {str(e)}"


class ExcelSetCellTool(Tool):
    """Tool to set cell value in Excel file."""

    def __init__(self, workspace: Path | None = None, allowed_dir: Path | None = None):
        self._workspace = workspace
        self._allowed_dir = allowed_dir

    @property
    def name(self) -> str:
        return "excel_set_cell"

    @property
    def description(self) -> str:
        return "Set value of a specific cell in Excel file."

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "The absolute path to the Excel file"
                },
                "sheet_name": {
                    "type": "string",
                    "description": "Sheet name. Default: first sheet"
                },
                "cell": {
                    "type": "string",
                    "description": "Cell reference, e.g., 'A1', 'B2'"
                },
                "value": {
                    "type": "string",
                    "description": "Value to set"
                }
            },
            "required": ["path", "cell", "value"]
        }

    async def execute(self, path: str, cell: str, value: str,
                     sheet_name: str | None = None, **kwargs: Any) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return "Error: openpyxl is required. Install with: pip install openpyxl"

        try:
            file_path = _resolve_path(path, self._workspace, self._allowed_dir)
            if not file_path.exists():
                return f"Error: File not found: {path}"

            wb = load_workbook(filename=str(file_path))

            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    wb.close()
                    return f"Error: Sheet '{sheet_name}' not found"
                ws = wb[sheet_name]
            else:
                ws = wb.active

            ws[cell] = value
            wb.save(str(file_path))
            wb.close()

            return f"Successfully set {cell} = '{value}' in {file_path}"

        except PermissionError as e:
            return f"Error: {e}"
        except Exception as e:
            return f"Error setting cell: {str(e)}"

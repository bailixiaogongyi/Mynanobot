"""Legacy document format conversion tools.

Supports reading and converting:
- .doc -> .docx (using textract or win32com)
- .xls -> .xlsx (using xlrd or win32com)
"""

import json
import platform
import subprocess
import shutil
from pathlib import Path
from typing import Any

from nanobot.agent.tools.base import Tool


def _resolve_path(path: str, workspace: Path | None = None, allowed_dir: Path | None = None) -> Path:
    """Resolve path against workspace and enforce directory restriction."""
    p = Path(path).expanduser()
    if not p.is_absolute() and workspace:
        p = workspace / p
    resolved = p.resolve()
    if allowed_dir:
        try:
            resolved.relative_to(allowed_dir.resolve())
        except ValueError:
            raise PermissionError(f"Path {path} is outside allowed directory {allowed_dir}")
    return resolved


class XlsReadTool(Tool):
    """Tool to read data from legacy .xls Excel file."""

    def __init__(self, workspace: Path | None = None, allowed_dir: Path | None = None):
        self._workspace = workspace
        self._allowed_dir = allowed_dir

    @property
    def name(self) -> str:
        return "xls_read"

    @property
    def description(self) -> str:
        return "Read data from legacy .xls Excel file (Excel 97-2003 format). Returns data in JSON format."

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "The absolute path to the .xls file"
                },
                "sheet_name": {
                    "type": "string",
                    "description": "Sheet name to read. Default: first sheet"
                },
                "header_row": {
                    "type": "boolean",
                    "description": "Whether first row is header. Default: true"
                },
                "max_rows": {
                    "type": "integer",
                    "description": "Maximum number of rows to read. Default: all"
                }
            },
            "required": ["path"]
        }

    async def execute(self, path: str, sheet_name: str | None = None,
                     header_row: bool = True, max_rows: int | None = None, **kwargs: Any) -> str:
        try:
            import xlrd
        except ImportError:
            return "Error: xlrd is required. Install with: pip install xlrd==1.2.0"

        try:
            file_path = _resolve_path(path, self._workspace, self._allowed_dir)
            if not file_path.exists():
                return f"Error: File not found: {path}"
            if file_path.suffix.lower() != ".xls":
                return f"Error: Not a .xls file: {path}"

            workbook = xlrd.open_workbook(str(file_path))

            if sheet_name:
                sheet = None
                for s in workbook.sheets():
                    if s.name == sheet_name:
                        sheet = s
                        break
                if sheet is None:
                    return f"Error: Sheet '{sheet_name}' not found. Available: {', '.join(workbook.sheet_names())}"
            else:
                sheet = workbook.sheet_by_index(0)

            rows = []
            for row_idx in range(sheet.nrows):
                row_data = []
                for col_idx in range(sheet.ncols):
                    cell = sheet.cell(row_idx, col_idx)
                    if cell.ctype == xlrd.XL_CELL_TEXT:
                        row_data.append(cell.value)
                    elif cell.ctype == xlrd.XL_CELL_NUMBER:
                        row_data.append(cell.value)
                    elif cell.ctype == xlrd.XL_CELL_DATE:
                        row_data.append(xlrd.xldate_as_datetime(cell.value, workbook.datemode).isoformat())
                    elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                        row_data.append(bool(cell.value))
                    else:
                        row_data.append(str(cell.value) if cell.value else "")
                rows.append(row_data)

            if max_rows:
                rows = rows[:max_rows]

            headers = []
            data = []

            if header_row and rows:
                headers = [str(cell) if cell else "" for cell in rows[0]]
                for row in rows[1:]:
                    row_dict = {}
                    for i, cell in enumerate(row):
                        key = headers[i] if i < len(headers) else f"col_{i}"
                        row_dict[key] = str(cell) if cell is not None else ""
                    data.append(row_dict)
            else:
                data = rows

            result = {
                "sheet_name": sheet.name,
                "total_rows": len(data),
                "headers": headers if header_row else None,
                "data": data
            }

            return json.dumps(result, ensure_ascii=False, indent=2)

        except PermissionError as e:
            return f"Error: {e}"
        except Exception as e:
            return f"Error reading .xls file: {str(e)}"


class XlsToXlsxTool(Tool):
    """Tool to convert .xls to .xlsx format."""

    def __init__(self, workspace: Path | None = None, allowed_dir: Path | None = None):
        self._workspace = workspace
        self._allowed_dir = allowed_dir

    @property
    def name(self) -> str:
        return "xls_to_xlsx"

    @property
    def description(self) -> str:
        return "Convert legacy .xls Excel file to modern .xlsx format."

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "The absolute path to the .xls file"
                },
                "output_path": {
                    "type": "string",
                    "description": "Output path for .xlsx file. Default: same directory with .xlsx extension"
                }
            },
            "required": ["path"]
        }

    async def execute(self, path: str, output_path: str | None = None, **kwargs: Any) -> str:
        try:
            import xlrd
            from openpyxl import Workbook
        except ImportError as e:
            return f"Error: Required libraries not installed. Install with: pip install xlrd==1.2.0 openpyxl"

        try:
            file_path = _resolve_path(path, self._workspace, self._allowed_dir)
            if not file_path.exists():
                return f"Error: File not found: {path}"
            if file_path.suffix.lower() != ".xls":
                return f"Error: Not a .xls file: {path}"

            out_path = _resolve_path(output_path, self._workspace, self._allowed_dir) if output_path else file_path.with_suffix(".xlsx")

            xls_workbook = xlrd.open_workbook(str(file_path))
            xlsx_workbook = Workbook()
            xlsx_workbook.remove(xlsx_workbook.active)

            for sheet_index in range(xls_workbook.nsheets):
                xls_sheet = xls_workbook.sheet_by_index(sheet_index)
                xlsx_sheet = xlsx_workbook.create_sheet(title=xls_sheet.name)

                for row_idx in range(xls_sheet.nrows):
                    for col_idx in range(xls_sheet.ncols):
                        cell = xls_sheet.cell(row_idx, col_idx)
                        if cell.ctype == xlrd.XL_CELL_TEXT:
                            xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=cell.value)
                        elif cell.ctype == xlrd.XL_CELL_NUMBER:
                            xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=cell.value)
                        elif cell.ctype == xlrd.XL_CELL_DATE:
                            xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1, 
                                          value=xlrd.xldate_as_datetime(cell.value, xls_workbook.datemode))
                        elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                            xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=bool(cell.value))
                        else:
                            xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=str(cell.value) if cell.value else "")

            out_path.parent.mkdir(parents=True, exist_ok=True)
            xlsx_workbook.save(str(out_path))

            return f"Successfully converted {file_path} to {out_path}"

        except PermissionError as e:
            return f"Error: {e}"
        except Exception as e:
            return f"Error converting .xls to .xlsx: {str(e)}"


class DocReadTool(Tool):
    """Tool to read text from legacy .doc Word file."""

    def __init__(self, workspace: Path | None = None, allowed_dir: Path | None = None):
        self._workspace = workspace
        self._allowed_dir = allowed_dir

    @property
    def name(self) -> str:
        return "doc_read"

    @property
    def description(self) -> str:
        return "读取旧版 .doc Word 文件 (Word 97-2003 格式) 的文本内容。当用户上传或询问 .doc 文件时使用此工具。"

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "The absolute path to the .doc file"
                }
            },
            "required": ["path"]
        }

    async def execute(self, path: str, **kwargs: Any) -> str:
        file_path = _resolve_path(path, self._workspace, self._allowed_dir)
        if not file_path.exists():
            return f"Error: File not found: {path}"
        if file_path.suffix.lower() != ".doc":
            return f"Error: Not a .doc file: {path}"

        if platform.system() == "Windows":
            return await self._read_with_win32com(file_path)
        else:
            return await self._read_with_textract(file_path)

    async def _read_with_win32com(self, file_path: Path) -> str:
        try:
            import win32com.client
        except ImportError:
            return "Error: pywin32 is required on Windows. Install with: pip install pywin32"

        try:
            word = win32com.client.Dispatch("Word.Application")
            word.Visible = False
            doc = word.Documents.Open(str(file_path))
            text = doc.Content.Text
            doc.Close()
            word.Quit()
            return text
        except Exception as e:
            return f"Error reading .doc with Word COM: {str(e)}"

    async def _read_with_textract(self, file_path: Path) -> str:
        try:
            import textract
        except ImportError:
            return "Error: textract is required. Install with: pip install textract"

        try:
            text = textract.process(str(file_path))
            return text.decode('utf-8')
        except Exception as e:
            return f"Error reading .doc with textract: {str(e)}"


class DocToDocxTool(Tool):
    """Tool to convert .doc to .docx format."""

    def __init__(self, workspace: Path | None = None, allowed_dir: Path | None = None):
        self._workspace = workspace
        self._allowed_dir = allowed_dir

    @property
    def name(self) -> str:
        return "doc_to_docx"

    @property
    def description(self) -> str:
        return "Convert legacy .doc Word file to modern .docx format."

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "The absolute path to the .doc file"
                },
                "output_path": {
                    "type": "string",
                    "description": "Output path for .docx file. Default: same directory with .docx extension"
                }
            },
            "required": ["path"]
        }

    async def execute(self, path: str, output_path: str | None = None, **kwargs: Any) -> str:
        file_path = _resolve_path(path, self._workspace, self._allowed_dir)
        if not file_path.exists():
            return f"Error: File not found: {path}"
        if file_path.suffix.lower() != ".doc":
            return f"Error: Not a .doc file: {path}"

        out_path = _resolve_path(output_path, self._workspace, self._allowed_dir) if output_path else file_path.with_suffix(".docx")

        if platform.system() == "Windows":
            return await self._convert_with_win32com(file_path, out_path)
        else:
            return await self._convert_with_libreoffice(file_path, out_path)

    async def _convert_with_win32com(self, file_path: Path, out_path: Path) -> str:
        try:
            import win32com.client
        except ImportError:
            return "Error: pywin32 is required on Windows. Install with: pip install pywin32"

        try:
            word = win32com.client.Dispatch("Word.Application")
            word.Visible = False
            doc = word.Documents.Open(str(file_path))
            out_path.parent.mkdir(parents=True, exist_ok=True)
            doc.SaveAs(str(out_path), FileFormat=16)
            doc.Close()
            word.Quit()
            return f"Successfully converted {file_path} to {out_path}"
        except Exception as e:
            return f"Error converting .doc with Word COM: {str(e)}"

    async def _convert_with_libreoffice(self, file_path: Path, out_path: Path) -> str:
        soffice = shutil.which("soffice") or shutil.which("libreoffice")
        if not soffice:
            return "Error: LibreOffice not found. Install LibreOffice to convert .doc files on Linux."

        try:
            out_path.parent.mkdir(parents=True, exist_ok=True)
            result = subprocess.run(
                [soffice, "--headless", "--convert-to", "docx", "--outdir", str(out_path.parent), str(file_path)],
                capture_output=True,
                text=True,
                timeout=60
            )
            if result.returncode == 0:
                return f"Successfully converted {file_path} to {out_path}"
            else:
                return f"Error converting with LibreOffice: {result.stderr}"
        except subprocess.TimeoutExpired:
            return "Error: LibreOffice conversion timed out"
        except Exception as e:
            return f"Error converting .doc with LibreOffice: {str(e)}"


LEGACY_TOOLS = [
    XlsReadTool,
    XlsToXlsxTool,
    DocReadTool,
    DocToDocxTool,
]
